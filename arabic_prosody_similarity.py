# -*- coding: utf-8 -*-
"""
Created on Sun Apr 22 20:38:46 2018

@author: EREN
"""

# -*- coding: utf-8 -*-
"""
Created on Sun Apr 22 13:09:17 2018

@author: EREN
"""
from builtins import str
from openpyxl import load_workbook
from difflib import SequenceMatcher

def similar(a, b):
    return SequenceMatcher(None, a, b).ratio()

def lowercase(ch):
    return {
    'İ':u'i',
    'I':u'ı',
    'Ç':u'ç',
    'Ğ':u'ğ',
    'Ş':u'ş'   
    }.get(ch, ch.lower())
             
def sesli(ch):
    ch = lowercase(ch)
    if ch in [u'a', u'e', u'i', u'ı', u'o', u'ö', u'u', u'ü']:
        return True
    else:
        return False
 
def hecele(str, heceler):
    index=0
    length=len(str)
    while sesli(str[index]) == False and length>index+1:
        index=index+1
    try:
        if sesli(str[index+1]):
            heceler.append(str[0:index+1])
            hecele(str[index+1:],heceler)
        elif length>index+2:
            if sesli(str[index+2]):
                heceler.append(str[0:index+1])
                hecele(str[index+1:],heceler)
            elif length>index+3:
                if sesli(str[index+3]):
                    heceler.append(str[0:index+2])
                    hecele(str[index+2:],heceler)
                else:
                    if str[index+1:index+4] in [u'str', u'ktr', u'mtr', u'nsp']:
                        #print "istisna!.."
                        heceler.append(str[0:index+2])
                        hecele(str[index+2:],heceler)
                    else:
                        #print "üç sessiz, normal kural"
                        heceler.append(str[0:index+3])
                        hecele(str[index+3:],heceler)
            else:
                heceler.append((str))
        else: 
            heceler.append((str))
    except:
        heceler.append((str))
        return
             
            
def aruzify_word(str):
    
    aruzified_word=""
    heceler = []
    hecele(str,heceler)
    
    for hece in heceler:
        hece= hece.lower()
        if len(hece)==1 and (hece=='a' or hece=='e' or hece=='ı' or hece=='i' or hece=='o' or hece=='ö' or hece=='u' or hece=='ü'):
            aruzified_word = aruzified_word + "."
    
        elif len(hece) == 2 and (hece[0]=='a' or hece[0]=='e' or hece[0]=='ı' or hece[0]=='i' or hece[0]=='o' or hece[0]=='ö' or hece[0]=='u' or hece[0]=='ü') and not (hece[1]=='a' or hece[1]=='e' or hece[1]=='ı' or hece[1]=='i' or hece[1]=='o' or hece[1]=='ö' or hece[1]=='u' or hece[1]=='ü'):
            aruzified_word = aruzified_word + "-"
            
        elif len(hece) == 2 and not (hece[0]=='a' or hece[0]=='e' or hece[0]=='ı' or hece[0]=='i' or hece[0]=='o' or hece[0]=='ö' or hece[0]=='u' or hece[0]=='ü') and (hece[1]=='a' or hece[1]=='e' or hece[1]=='ı' or hece[1]=='i' or hece[1]=='o' or hece[1]=='ö' or hece[1]=='u' or hece[1]=='ü'):
            aruzified_word = aruzified_word + "."
            
        elif len(hece) == 3 and not (hece[0]=='a' or hece[0]=='e' or hece[0]=='ı' or hece[0]=='i' or hece[0]=='o' or hece[0]=='ö' or hece[0]=='u' or hece[0]=='ü') and (hece[1]=='a' or hece[1]=='e' or hece[1]=='ı' or hece[1]=='i' or hece[1]=='o' or hece[1]=='ö' or hece[1]=='u' or hece[1]=='ü') and not (hece[2]=='a' or hece[2]=='e' or hece[2]=='ı' or hece[2]=='i' or hece[2]=='o' or hece[2]=='ö' or hece[2]=='u' or hece[2]=='ü'):
            aruzified_word = aruzified_word + "-"
            
        elif len(hece) == 3 and (hece[0]=='a' or hece[0]=='e' or hece[0]=='ı' or hece[0]=='i' or hece[0]=='o' or hece[0]=='ö' or hece[0]=='u' or hece[0]=='ü') and not (hece[1]=='a' or hece[1]=='e' or hece[1]=='ı' or hece[1]=='i' or hece[1]=='o' or hece[1]=='ö' or hece[1]=='u' or hece[1]=='ü') and not (hece[2]=='a' or hece[2]=='e' or hece[2]=='ı' or hece[2]=='i' or hece[2]=='o' or hece[2]=='ö' or hece[2]=='u' or hece[2]=='ü'):
            aruzified_word = aruzified_word + "-"          
        
        elif len(hece) == 4 and not (hece[0]=='a' or hece[0]=='e' or hece[0]=='ı' or hece[0]=='i' or hece[0]=='o' or hece[0]=='ö' or hece[0]=='u' or hece[0]=='ü') and (hece[1]=='a' or hece[1]=='e' or hece[1]=='ı' or hece[1]=='i' or hece[1]=='o' or hece[1]=='ö' or hece[1]=='u' or hece[1]=='ü') and not (hece[2]=='a' or hece[2]=='e' or hece[2]=='ı' or hece[2]=='i' or hece[2]=='o' or hece[2]=='ö' or hece[2]=='u' or hece[2]=='ü') and not (hece[3]=='a' or hece[3]=='e' or hece[3]=='ı' or hece[3]=='i' or hece[3]=='o' or hece[3]=='ö' or hece[3]=='u' or hece[3]=='ü'):
            aruzified_word = aruzified_word + "-"   
            
    return aruzified_word
            
def aruzify_poem(poem):
    poemStr=[]
    aruz_line=""
    count=0
    for line in poem.split('\n'):
        aruz_line=""
        for word in line.split():
            aruz_line = aruz_line + aruzify_word(word)
        poemStr.append(aruz_line)
        count = count + 1
        if(count == 2):
            break
        
    return poemStr

aruz_patterns = [".–.–.–.–","––.–––.–––.–––.–","––.––––.––––.––––.––",".––.––.––.––",".–.-.–.–.––","..--..--..--..–","–.–––.–––.–––.–","–.–––.–––.–","–..-–..––.–",".––.––.––.–",".–.–.––.–.–.––","..—.–.–..–",".–.––.–.––","––.–.–..––.–.–","––..–––.––","––..––..––..––","––..–––––..––","––..––..––","–..––.––..––.–"]

def aruz_skor(poemStr, aruz_patterns):
    
    max_similarity=0
    for pattern in aruz_patterns:
        sim_sum=0
        for dize in poemStr:    
            sim_sum = sim_sum + similar(dize,pattern)
        if((sim_sum)/len(poemStr)) > max_similarity:
            max_similarity = (sim_sum)/len(poemStr)

    return max_similarity

def get_aruz_olcu_skor(poem):        
    return aruz_skor(aruzify_poem(poem),aruz_patterns)

def get_sim(poem1,poem2):   
    poem1_aruz_skor = get_aruz_olcu_skor(poem1)
    poem2_aruz_skor = get_aruz_olcu_skor(poem2)
    
    distance = abs((poem1_aruz_skor - poem2_aruz_skor))
    
    if 1 - distance < 0:
        return 0 
    else:
        return 1 - distance
    
wb = load_workbook('allPoems_sonn.xlsx')
ws = wb.active
countt=0
sim_matrix = [[0 for i in range(2827)] for j in range(2827)]
file = open("aruz_sim_son.txt","w")
for i in range(1,2828):
    print(countt)
    cx1 = "C"+str(i)
    poem1 = ws[cx1]
    list = []
    for j in range(i,2828):
        cx2 = "C"+str(j)
        poem2 = ws[cx2]
        sim_matrix[i-1][j-1]=(get_sim(poem1.value,poem2.value))
    countt = countt+1
   
for i in range(0,2827):
    for j in range(0,i):
        sim_matrix[i][j] = sim_matrix[j][i]
             
for i in range(0,2827):
    print(i,"matrix")
    for j in range(0,2827):
        file.write(str(sim_matrix[i][j]) + " ")        
    file.write("\n")
