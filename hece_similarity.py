# -*- coding: utf-8 -*-
"""
Created on Sun Apr 22 20:38:46 2018

@author: EREN
"""

# -*- coding: utf-8 -*-
"""
Created on Sun Apr 22 13:09:17 2018

@author: EREN
"""
import numpy
from builtins import str
from openpyxl import load_workbook

def get_hece_olcu_skor(poem): 
    list = []
    if poem is not None:
        for line in poem.split('\n'):
            count = 0
            for i in line:
                if i.lower()=='a' or i.lower()=='e' or i.lower()=='ı' or i.lower()=='i' or i.lower()=='o' or i.lower()=='ö' or i.lower()=='u' or i.lower()=='ü':
                    count = count+1
            if count !=0:
                list.append(count)
            
        return 1 - (numpy.std(list)/(sum(list)/len(list)))
        
    else:
        return "no-score"

def get_sim(poem1,poem2):   
    poem1_hece_skor = get_hece_olcu_skor(poem1)    
    poem2_hece_skor = get_hece_olcu_skor(poem2)

    distance = abs((poem1_hece_skor - poem2_hece_skor))
    
    if 1 - distance < 0:
        return 0 
    else:
        return 1 - distance

wb = load_workbook('allPoems_sonn.xlsx')
ws = wb.active
countt=0
sim_matrix = [[0 for i in range(2827)] for j in range(2827)]
file = open("olcu_sim_son.txt","w")

for i in range(1,2828):
    print(countt)
    cx1 = "C"+str(i)
    poem1 = ws[cx1]
    list = []
    for j in range(i,2828):
        cx2 = "C"+str(j)
        poem2 = ws[cx2]
        sim_matrix[i-1][j-1]=(get_sim(poem1.value,poem2.value))
    countt = countt+1
   
for i in range(0,2827):
    for j in range(0,i):
        sim_matrix[i][j] = sim_matrix[j][i]
             
for i in range(0,2827):
    print(i,"matrix")
    for j in range(0,2827):
        file.write(str(sim_matrix[i][j]) + " ")        
    file.write("\n")

